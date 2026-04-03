#!/usr/bin/env python3
"""
Build a minimal standalone DCF workbook with live Excel formulas.
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BLUE_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="DDEBF7")
BOLD = Font(bold=True)
WHITE_BOLD = Font(bold=True, color="FFFFFF")


def pct_list(text: str) -> list[float]:
    return [float(part.strip()) for part in text.split(",") if part.strip()]


def set_input(ws, cell: str, value, source: str) -> None:
    ws[cell] = value
    ws[cell].fill = BLUE_FILL
    ws[cell].comment = Comment(source, "AIGroup")


def set_pct(ws, cell: str, value, source: str) -> None:
    set_input(ws, cell, value, source)
    ws[cell].number_format = "0.0%"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", required=True)
    parser.add_argument("--company", default="Apple Inc.")
    parser.add_argument("--ticker", default="AAPL")
    parser.add_argument("--currency", default="USD")
    parser.add_argument("--fiscal-year", type=int, default=2025)
    parser.add_argument("--base-revenue", type=float, required=True)
    parser.add_argument("--base-ebit", type=float, required=True)
    parser.add_argument("--shares", type=float, required=True)
    parser.add_argument("--net-debt", type=float, required=True)
    parser.add_argument("--price", type=float, required=True)
    parser.add_argument("--tax-rate", type=float, required=True)
    parser.add_argument("--wacc", type=float, required=True)
    parser.add_argument("--terminal-growth", type=float, required=True)
    parser.add_argument("--da-pct", type=float, required=True)
    parser.add_argument("--capex-pct", type=float, required=True)
    parser.add_argument("--nwc-pct", type=float, required=True)
    parser.add_argument("--revenue-growth", required=True)
    parser.add_argument("--ebit-margin", required=True)
    args = parser.parse_args()

    revenue_growth = pct_list(args.revenue_growth)
    ebit_margin = pct_list(args.ebit_margin)
    if len(revenue_growth) != 5 or len(ebit_margin) != 5:
        raise SystemExit("revenue-growth and ebit-margin must each contain 5 comma-separated values")

    wb = Workbook()
    ws = wb.active
    ws.title = "DCF"
    wacc_ws = wb.create_sheet("WACC")
    sens_ws = wb.create_sheet("Sensitivity")

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws["A1"] = f"{args.company} ({args.ticker}) - Minimal DCF"
    ws["A1"].font = Font(size=14, bold=True)

    ws["A3"] = "Market Data & Inputs"
    ws["A3"].fill = HEADER_FILL
    ws["A3"].font = WHITE_BOLD
    ws.merge_cells("A3:D3")

    labels = [
        ("A4", "Current price", "B4", args.price),
        ("A5", "Base revenue", "B5", args.base_revenue),
        ("A6", "Base EBIT", "B6", args.base_ebit),
        ("A7", "Diluted shares", "B7", args.shares),
        ("A8", "Net debt", "B8", args.net_debt),
        ("A9", "Tax rate", "B9", args.tax_rate),
        ("A10", "WACC", "B10", args.wacc),
        ("A11", "Terminal growth", "B11", args.terminal_growth),
        ("A12", "D&A % rev", "B12", args.da_pct),
        ("A13", "CapEx % rev", "B13", args.capex_pct),
        ("A14", "NWC % delta rev", "B14", args.nwc_pct),
    ]
    for lcell, label, vcell, value in labels:
        ws[lcell] = label
        set_input(ws, vcell, value, f"Source: user/MCP smoke test input, FY{args.fiscal_year}")
    for cell in ("B9", "B10", "B11", "B12", "B13", "B14"):
        ws[cell].number_format = "0.0%"

    start_col = 4
    hist_col = 3
    header_row = 17
    ws["A17"] = "DCF Projection"
    ws["A17"].fill = HEADER_FILL
    ws["A17"].font = WHITE_BOLD
    ws.merge_cells("A17:H17")

    ws["A19"] = "Metric"
    ws["C19"] = f"FY{args.fiscal_year}"
    for i in range(5):
        ws.cell(header_row + 2, start_col + i, args.fiscal_year + i + 1)
    years_row = 19
    for i in range(5):
        cell = ws.cell(years_row, start_col + i)
        cell.value = args.fiscal_year + i + 1
        cell.font = BOLD
        cell.fill = SECTION_FILL

    rows = {
        "revenue_growth": 20,
        "revenue": 21,
        "ebit_margin": 22,
        "ebit": 23,
        "tax_rate": 24,
        "nopat": 25,
        "da_pct": 26,
        "da": 27,
        "capex_pct": 28,
        "capex": 29,
        "nwc_pct": 30,
        "delta_nwc": 31,
        "ufcf": 32,
        "discount_factor": 33,
        "pv_ufcf": 34,
    }
    labels2 = {
        "revenue_growth": "Revenue growth",
        "revenue": "Revenue",
        "ebit_margin": "EBIT margin",
        "ebit": "EBIT",
        "tax_rate": "Tax rate",
        "nopat": "NOPAT",
        "da_pct": "D&A % revenue",
        "da": "D&A",
        "capex_pct": "CapEx % revenue",
        "capex": "CapEx",
        "nwc_pct": "NWC % delta revenue",
        "delta_nwc": "Delta NWC",
        "ufcf": "Unlevered FCF",
        "discount_factor": "Discount factor",
        "pv_ufcf": "PV of UFCF",
    }
    for key, row in rows.items():
        ws[f"A{row}"] = labels2[key]

    ws["C21"] = args.base_revenue
    ws["C23"] = args.base_ebit
    ws["C22"] = "=C23/C21"
    ws["C24"] = "=B9"
    ws["C25"] = "=C23*(1-C24)"
    ws["C26"] = "=B12"
    ws["C27"] = "=C21*C26"
    ws["C28"] = "=B13"
    ws["C29"] = "=C21*C28"
    ws["C30"] = "=B14"
    ws["C31"] = "0"
    ws["C32"] = "=C25+C27-C29-C31"
    ws["C33"] = "1"
    ws["C34"] = "=C32*C33"

    for i in range(5):
        col = get_column_letter(start_col + i)
        prev_col = get_column_letter(start_col + i - 1 if i > 0 else hist_col)
        growth_cell = f"{col}{rows['revenue_growth']}"
        margin_cell = f"{col}{rows['ebit_margin']}"
        set_pct(ws, growth_cell, revenue_growth[i], f"Source: smoke test conservative assumption for {args.ticker}")
        set_pct(ws, margin_cell, ebit_margin[i], f"Source: smoke test conservative assumption for {args.ticker}")
        ws[f"{col}{rows['revenue']}"] = f"={prev_col}{rows['revenue']}*(1+{growth_cell})"
        ws[f"{col}{rows['ebit']}"] = f"={col}{rows['revenue']}*{margin_cell}"
        ws[f"{col}{rows['tax_rate']}"] = "=B9"
        ws[f"{col}{rows['nopat']}"] = f"={col}{rows['ebit']}*(1-{col}{rows['tax_rate']})"
        ws[f"{col}{rows['da_pct']}"] = "=B12"
        ws[f"{col}{rows['da']}"] = f"={col}{rows['revenue']}*{col}{rows['da_pct']}"
        ws[f"{col}{rows['capex_pct']}"] = "=B13"
        ws[f"{col}{rows['capex']}"] = f"={col}{rows['revenue']}*{col}{rows['capex_pct']}"
        ws[f"{col}{rows['nwc_pct']}"] = "=B14"
        ws[f'{col}{rows["delta_nwc"]}'] = f"=({col}{rows['revenue']}-{prev_col}{rows['revenue']})*{col}{rows['nwc_pct']}"
        ws[f"{col}{rows['ufcf']}"] = f"={col}{rows['nopat']}+{col}{rows['da']}-{col}{rows['capex']}-{col}{rows['delta_nwc']}"
        ws[f"{col}{rows['discount_factor']}"] = f"=1/(1+$B$10)^({i+1})"
        ws[f"{col}{rows['pv_ufcf']}"] = f"={col}{rows['ufcf']}*{col}{rows['discount_factor']}"

    tv_row = 37
    ws[f"A{tv_row}"] = "Terminal value"
    ws[f"B{tv_row}"] = "=H32*(1+$B$11)/($B$10-$B$11)"
    ws[f"A{tv_row+1}"] = "PV terminal value"
    ws[f"B{tv_row+1}"] = "=B37*H33"
    ws[f"A{tv_row+2}"] = "Enterprise value"
    ws[f"B{tv_row+2}"] = "=SUM(C34:H34)+B38"
    ws[f"A{tv_row+3}"] = "Equity value"
    ws[f"B{tv_row+3}"] = "=B39-B8"
    ws[f"A{tv_row+4}"] = "Implied value / share"
    ws[f"B{tv_row+4}"] = "=B40/B7"
    ws[f"A{tv_row+5}"] = "Upside / downside"
    ws[f"B{tv_row+5}"] = "=B41/B4-1"
    ws["B41"].number_format = "$#,##0.00"
    ws["B42"].number_format = "0.0%"

    wacc_ws["A1"] = "WACC"
    wacc_ws["A1"].fill = HEADER_FILL
    wacc_ws["A1"].font = WHITE_BOLD
    wacc_ws["A3"] = "WACC"
    wacc_ws["B3"] = "=DCF!B10"
    wacc_ws["A4"] = "Terminal growth"
    wacc_ws["B4"] = "=DCF!B11"
    wacc_ws["B3"].number_format = "0.0%"
    wacc_ws["B4"].number_format = "0.0%"

    sens_ws["A1"] = "Sensitivity"
    sens_ws["A1"].fill = HEADER_FILL
    sens_ws["A1"].font = WHITE_BOLD
    sens_ws["A3"] = "WACC \\ g"
    wacc_points = [args.wacc - 0.02, args.wacc - 0.01, args.wacc, args.wacc + 0.01, args.wacc + 0.02]
    g_points = [
        args.terminal_growth - 0.01,
        args.terminal_growth - 0.005,
        args.terminal_growth,
        args.terminal_growth + 0.005,
        args.terminal_growth + 0.01,
    ]
    for idx, val in enumerate(g_points, start=2):
        sens_ws.cell(3, idx, val).number_format = "0.0%"
    for idx, val in enumerate(wacc_points, start=4):
        sens_ws.cell(idx, 1, val).number_format = "0.0%"
    for r_idx in range(5):
        for c_idx in range(5):
            cell = sens_ws.cell(4 + r_idx, 2 + c_idx)
            wacc_ref = sens_ws.cell(4 + r_idx, 1).coordinate
            g_ref = sens_ws.cell(3, 2 + c_idx).coordinate
            cell.value = (
                f"=((DCF!H32*(1+{g_ref})/({wacc_ref}-{g_ref}))*DCF!H33+SUM(DCF!C34:H34)-DCF!B8)/DCF!B7"
            )
            cell.number_format = "$#,##0.00"
    sens_ws["D6"].fill = BLUE_FILL
    sens_ws["D6"].font = BOLD

    for ws_ in (ws, wacc_ws, sens_ws):
        for row in ws_.iter_rows():
            for cell in row:
                if cell.coordinate.startswith("B") and cell.row in (4, 5, 6, 7, 8):
                    cell.number_format = "$#,##0.00"
        ws_.freeze_panes = "B4"

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


if __name__ == "__main__":
    main()
