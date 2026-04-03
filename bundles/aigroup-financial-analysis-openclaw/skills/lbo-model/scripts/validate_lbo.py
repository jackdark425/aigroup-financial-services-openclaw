#!/usr/bin/env python3
"""
Validate a minimal LBO workbook for basic integrity.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl


REQUIRED_SHEETS = {"LBO"}
REQUIRED_FORMULAS = {"B6", "B7", "B18", "B19", "H20", "H21", "H22"}


def main() -> None:
    if len(sys.argv) < 2:
        print(json.dumps({"status": "error", "message": "missing path"}))
        raise SystemExit(1)

    path = Path(sys.argv[1])
    if not path.exists():
        print(json.dumps({"status": "error", "message": "file not found"}))
        raise SystemExit(1)

    wb = openpyxl.load_workbook(path, data_only=False)
    missing_sheets = sorted(REQUIRED_SHEETS - set(wb.sheetnames))
    if missing_sheets:
        print(json.dumps({"status": "fail", "missingSheets": missing_sheets}))
        raise SystemExit(1)

    ws = wb["LBO"]
    missing_formulas = []
    for cell in REQUIRED_FORMULAS:
        value = ws[cell].value
        if not (isinstance(value, str) and value.startswith("=")):
            missing_formulas.append(cell)

    result = {
        "status": "PASS" if not missing_formulas else "FAIL",
        "file": str(path),
        "missingFormulas": missing_formulas,
        "sheets": wb.sheetnames,
        "momFormula": ws["H22"].value,
        "entryDebtFormula": ws["B6"].value,
        "equityFormula": ws["B7"].value,
    }
    print(json.dumps(result))
    if missing_formulas:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
