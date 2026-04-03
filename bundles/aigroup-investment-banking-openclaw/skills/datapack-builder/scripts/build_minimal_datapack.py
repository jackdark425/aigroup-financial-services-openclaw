#!/usr/bin/env python3
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


HEADER_FILL = PatternFill("solid", fgColor="4472C4")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")


def write_header(ws, row: int, text: str) -> None:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = HEADER_FILL


def write_label(ws, row: int, col: int, text: str, bold: bool = False) -> None:
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=bold)


def build_workbook(company: str, revenue: float, ebitda_margin: float, vertical: str, geography: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Summary"
    write_header(ws, 1, "Executive Summary")
    rows = [
        ("Company", company),
        ("Vertical", vertical),
        ("Geography", geography),
        ("Revenue ($mm)", revenue),
        ("EBITDA Margin", ebitda_margin),
        ("EBITDA ($mm)", "=B4*B5"),
        ("Growth Case", "Mid-teens expansion from enterprise upsell"),
        ("Sources", "User-provided fictional facts"),
    ]
    for idx, (k, v) in enumerate(rows, start=3):
        write_label(ws, idx, 1, k, bold=True)
        ws.cell(row=idx, column=2, value=v)
    ws["B5"].number_format = "0.0%"
    ws["B4"].number_format = "$#,##0.0"
    ws["B6"].number_format = "$#,##0.0"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 34

    hist = wb.create_sheet("Historical Financials")
    write_header(hist, 1, "Historical Financials")
    years = ["2023A", "2024A", "2025A"]
    metrics = [
        ("Revenue", [revenue * 0.78, revenue * 0.89, revenue]),
        ("EBITDA", [revenue * 0.78 * (ebitda_margin - 0.02), revenue * 0.89 * (ebitda_margin - 0.01), revenue * ebitda_margin]),
        ("EBITDA Margin", [(ebitda_margin - 0.02), (ebitda_margin - 0.01), ebitda_margin]),
    ]
    for c, year in enumerate(years, start=2):
        cell = hist.cell(row=3, column=c, value=year)
        cell.font = Font(bold=True)
        cell.fill = SUB_FILL
    for r, (metric, values) in enumerate(metrics, start=4):
        write_label(hist, r, 1, metric, bold=True)
        for c, value in enumerate(values, start=2):
            hist.cell(row=r, column=c, value=value)
            if "Margin" in metric:
                hist.cell(row=r, column=c).number_format = "0.0%"
            else:
                hist.cell(row=r, column=c).number_format = "$#,##0.0"

    ops = wb.create_sheet("Operating Metrics")
    write_header(ops, 1, "Operating Metrics")
    op_rows = [
        ("ARR Retention", 1.12),
        ("Gross Retention", 0.93),
        ("Enterprise Mix", 0.68),
        ("Customers", 420),
        ("Revenue per Customer ($k)", 107.1),
    ]
    for idx, (metric, value) in enumerate(op_rows, start=3):
        write_label(ops, idx, 1, metric, bold=True)
        ops.cell(row=idx, column=2, value=value)
        if isinstance(value, float) and value <= 2:
            ops.cell(row=idx, column=2).number_format = "0.0%"

    market = wb.create_sheet("Market Analysis")
    write_header(market, 1, "Market Analysis")
    market_rows = [
        ("Theme", "Vertical SaaS vendors with workflow lock-in"),
        ("Primary Buyers", "Mid-market and enterprise operators"),
        ("Key Risks", "Longer enterprise sales cycles; platform competition"),
        ("Key Catalysts", "Cross-sell analytics and international expansion"),
    ]
    for idx, (metric, value) in enumerate(market_rows, start=3):
        write_label(market, idx, 1, metric, bold=True)
        market.cell(row=idx, column=2, value=value)

    highlights = wb.create_sheet("Investment Highlights")
    write_header(highlights, 1, "Investment Highlights")
    bullets = [
        "Mission-critical vertical SaaS platform with recurring revenue model.",
        "Healthy profitability profile with EBITDA margin already above 20%.",
        "Attractive expansion path via pricing, add-on modules, and new logos.",
        "Compact datapack generated from fictional facts for smoke testing.",
    ]
    for idx, bullet in enumerate(bullets, start=3):
        highlights.cell(row=idx, column=1, value=f"- {bullet}")

    return wb


def build_summary(path: Path, company: str, revenue: float, ebitda_margin: float) -> None:
    path.write_text(
        "\n".join(
            [
                f"# {company} Datapack",
                "",
                "## Snapshot",
                f"- Revenue: ${revenue:.1f}mm",
                f"- EBITDA Margin: {ebitda_margin:.1%}",
                f"- Workbook: structured into executive summary, historicals, operating metrics, market analysis, and highlights.",
                "",
                "## Notes",
                "- Figures are fictional and intended for OpenClaw smoke testing.",
            ]
        )
        + "\n",
        encoding="utf-8",
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--company", required=True)
    parser.add_argument("--revenue", required=True, type=float)
    parser.add_argument("--ebitda-margin", required=True, type=float)
    parser.add_argument("--vertical", required=True)
    parser.add_argument("--geography", required=True)
    parser.add_argument("--xlsx-out", required=True)
    parser.add_argument("--summary-out", required=True)
    args = parser.parse_args()

    xlsx_out = Path(args.xlsx_out)
    summary_out = Path(args.summary_out)
    xlsx_out.parent.mkdir(parents=True, exist_ok=True)
    summary_out.parent.mkdir(parents=True, exist_ok=True)

    wb = build_workbook(args.company, args.revenue, args.ebitda_margin, args.vertical, args.geography)
    wb.save(xlsx_out)
    build_summary(summary_out, args.company, args.revenue, args.ebitda_margin)


if __name__ == "__main__":
    main()
