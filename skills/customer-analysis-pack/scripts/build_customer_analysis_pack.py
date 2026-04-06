#!/usr/bin/env python3
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill


HEADER_FILL = PatternFill("solid", fgColor="4472C4")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
INPUT_FILL = PatternFill("solid", fgColor="E2EFDA")


def split_items(raw: str | None, fallback: list[str]) -> list[str]:
    if not raw:
        return fallback
    values = [item.strip() for item in raw.split("|") if item.strip()]
    return values or fallback


def parse_fit_matrix(raw: str | None, fallback: list[str]) -> list[tuple[str, str, str, str]]:
    base = raw or "||".join(fallback)
    row_candidates = [item.strip() for item in base.split("||") if item.strip()]

    if len(row_candidates) == 1 and "|" in row_candidates[0]:
        flat = [item.strip() for item in row_candidates[0].split("|") if item.strip()]
        if len(flat) >= 4 and len(flat) % 4 == 0:
            row_candidates = ["|".join(flat[i : i + 4]) for i in range(0, len(flat), 4)]

    matrix: list[tuple[str, str, str, str]] = []
    for item in row_candidates:
        parts = [part.strip() for part in item.split("|")]
        while len(parts) < 4:
            parts.append("")
        matrix.append((parts[0], parts[1], parts[2], parts[3]))
    return matrix


def header(ws, row: int, text: str) -> None:
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = HEADER_FILL


def label(ws, row: int, col: int, text: str, bold: bool = False) -> None:
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=bold)


def input_cell(ws, row: int, col: int, value, number_format: str | None = None) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.fill = INPUT_FILL
    c.font = Font(color="0000FF")
    if number_format:
        c.number_format = number_format


def add_fit_chart(ws) -> None:
    chart = BarChart()
    chart.title = "产品匹配度"
    data = Reference(ws, min_col=2, max_col=2, min_row=3, max_row=7)
    cats = Reference(ws, min_col=1, max_col=1, min_row=3, max_row=7)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.style = 10
    chart.height = 7
    chart.width = 11
    ws.add_chart(chart, "D3")


def fit_score(label_text: str) -> int:
    mapping = {"高": 5, "中高": 4, "中": 3, "中低": 2, "低": 1}
    return mapping.get(label_text, 3)


def build_workbook(
    company: str,
    revenue: float,
    ebitda_margin: float,
    summary: str,
    fit_matrix: list[tuple[str, str, str, str]],
    next_actions: list[tuple[str, str, str, str]],
) -> Workbook:
    wb = Workbook()

    ws = wb.active
    ws.title = "客户概况"
    header(ws, 1, "客户概况")
    rows = [
        ("客户名称", company),
        ("收入（人民币百万元）", revenue),
        ("EBITDA 利润率", ebitda_margin),
        ("概要判断", summary),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        label(ws, i, 1, k, bold=True)
        fmt = "$#,##0.0" if i == 4 else "0.0%" if i == 5 else None
        input_cell(ws, i, 2, v, fmt)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 100

    fit = wb.create_sheet("产品匹配矩阵")
    header(fit, 1, "银行产品匹配矩阵")
    columns = ["产品", "匹配度分值", "匹配度", "理由", "触发条件"]
    for col, name in enumerate(columns, start=1):
        cell = fit.cell(row=2, column=col, value=name)
        cell.fill = SUB_FILL
        cell.font = Font(bold=True)
    for row, (product, level, reason, trigger) in enumerate(fit_matrix, start=3):
        fit.cell(row=row, column=1, value=product)
        fit.cell(row=row, column=2, value=fit_score(level))
        fit.cell(row=row, column=3, value=level)
        fit.cell(row=row, column=4, value=reason)
        fit.cell(row=row, column=5, value=trigger)
    fit.column_dimensions["A"].width = 18
    fit.column_dimensions["B"].width = 12
    fit.column_dimensions["C"].width = 10
    fit.column_dimensions["D"].width = 52
    fit.column_dimensions["E"].width = 28
    add_fit_chart(fit)

    risk = wb.create_sheet("风险判断")
    header(risk, 1, "风险判断")
    risk_rows = [
        ("财务风险", "中高", "需继续核实现金流、盈利能力和授信准入条件"),
        ("经营风险", "中", "客户集中度、产品迭代与商业化节奏仍需跟踪"),
        ("合规风险", "低", "若司法和经营异常记录较少，可作为相对正面信号"),
        ("行业风险", "中高", "受政策、供应链和竞争格局影响较大"),
    ]
    for col, name in enumerate(["类别", "等级", "说明"], start=1):
        cell = risk.cell(row=2, column=col, value=name)
        cell.fill = SUB_FILL
        cell.font = Font(bold=True)
    for row, values in enumerate(risk_rows, start=3):
        for col, value in enumerate(values, start=1):
            risk.cell(row=row, column=col, value=value)
    risk.column_dimensions["A"].width = 16
    risk.column_dimensions["B"].width = 10
    risk.column_dimensions["C"].width = 70

    actions = wb.create_sheet("下一步动作")
    header(actions, 1, "下一步动作")
    for col, name in enumerate(["动作", "负责人", "优先级", "说明"], start=1):
        cell = actions.cell(row=2, column=col, value=name)
        cell.fill = SUB_FILL
        cell.font = Font(bold=True)
    for row, values in enumerate(next_actions, start=3):
        for col, value in enumerate(values, start=1):
            actions.cell(row=row, column=col, value=value)
    actions.column_dimensions["A"].width = 30
    actions.column_dimensions["B"].width = 12
    actions.column_dimensions["C"].width = 10
    actions.column_dimensions["D"].width = 60

    return wb


def build_markdown(
    company: str,
    revenue: float,
    ebitda_margin: float,
    summary: str,
    fit_matrix: list[tuple[str, str, str, str]],
    risks: list[str],
    industry: list[str],
    next_steps: list[str],
) -> str:
    lines = [
        f"# {company}客户分析包",
        "",
        "> **内部使用 · 初步分析稿**",
        "> 本分析包基于客户调查结果与初步假设生成，正式使用前需进一步核验公开披露和审计材料。",
        "",
        "## 一、客户概况",
        "",
        summary,
        "",
        "## 二、财务快照",
        "",
        f"- 收入（参考）：人民币 {revenue:.1f} 百万元",
        f"- EBITDA 利润率（参考）：{ebitda_margin:.1%}",
        f"- EBITDA（参考）：人民币 {revenue * ebitda_margin:.1f} 百万元",
        "",
        "## 三、银行产品匹配矩阵",
        "",
        "| 产品 | 匹配度 | 理由 | 触发条件 |",
        "|------|--------|------|----------|",
    ]
    for product, level, reason, trigger in fit_matrix:
        lines.append(f"| {product} | {level} | {reason} | {trigger} |")
    lines += [
        "",
        "## 四、风险判断",
        "",
    ]
    lines.extend([f"- {item}" for item in risks])
    lines += [
        "",
        "## 五、行业与竞争观察",
        "",
    ]
    lines.extend([f"- {item}" for item in industry])
    lines += [
        "",
        "## 六、客户经理下一步动作",
        "",
    ]
    lines.extend([f"{idx + 1}. {item}" for idx, item in enumerate(next_steps)])
    lines += [
        "",
        "---",
        "",
        "*本文件为自动生成的客户分析包，可作为客户经理内部讨论、覆盖推进和后续 datapack / 模型工作的输入。*",
    ]
    return "\n".join(lines) + "\n"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--company", required=True)
    parser.add_argument("--revenue", type=float, default=800.0)
    parser.add_argument("--ebitda-margin", type=float, default=0.12)
    parser.add_argument("--summary", default="")
    parser.add_argument("--fit-matrix", default="")
    parser.add_argument("--risks", default="")
    parser.add_argument("--industry-observations", default="")
    parser.add_argument("--next-steps", default="")
    parser.add_argument("--xlsx-out", required=True)
    parser.add_argument("--md-out", required=True)
    args = parser.parse_args()

    summary = args.summary or (
        f"{args.company} 已完成初步客户调查，当前适合进入银行内部分析阶段，重点评估产品切入点、财务可核实性和风险边界。"
    )
    fit_matrix = parse_fit_matrix(
        args.fit_matrix,
        [
            "现金管理|高|资金沉淀和流动性管理空间明确|建立基础账户关系后",
            "供应链金融|高|上下游应收应付链条具备融资切口|获取核心合同和账期信息后",
            "项目融资|中高|若存在扩产、算力建设或重大项目，可衔接融资需求|项目立项或中标后",
            "授信|中|需先核实财务和担保条件|财报和主体资质明确后",
            "投行服务|中|若存在融资、并购或资本运作诉求，可跟进|资本动作出现后",
        ],
    )

    risks = split_items(
        args.risks,
        [
            "财务风险：收入、盈利和现金流仍需结合审计材料进一步核验。",
            "经营风险：客户集中度、产品迭代速度和订单可持续性需持续跟踪。",
            "行业风险：受政策、地缘政治、供应链和竞争格局影响较大。",
        ],
    )
    industry = split_items(
        args.industry_observations,
        [
            "行业景气与政策支持度较高，但竞争也更快集中到头部玩家。",
            "同业比较应重点关注技术路线、客户结构、盈利能力和生态能力。",
            "若处于国产替代或战略新兴赛道，银行覆盖价值通常高于传统行业平均水平。",
        ],
    )
    next_steps = split_items(
        args.next_steps,
        [
            "补齐审计财报、年报或季报中的核心财务字段。",
            "确认银行优先切入产品及对应触发条件。",
            "安排首次拜访或内部讨论，并形成后续覆盖计划。",
        ],
    )

    next_action_rows = []
    for idx, item in enumerate(next_steps):
        priority = "高" if idx == 0 else "中"
        next_action_rows.append((item, "RM", priority, "客户经理内部推进事项"))

    wb = build_workbook(
        args.company,
        args.revenue,
        args.ebitda_margin,
        summary,
        fit_matrix,
        next_action_rows,
    )
    xlsx_out = Path(args.xlsx_out)
    md_out = Path(args.md_out)
    xlsx_out.parent.mkdir(parents=True, exist_ok=True)
    md_out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_out)
    md_out.write_text(
        build_markdown(
            args.company,
            args.revenue,
            args.ebitda_margin,
            summary,
            fit_matrix,
            risks,
            industry,
            next_steps,
        ),
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
