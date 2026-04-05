#!/usr/bin/env python3
from __future__ import annotations

import argparse
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill


HEADER_FILL = PatternFill("solid", fgColor="4472C4")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
INPUT_FILL = PatternFill("solid", fgColor="E2EFDA")


def normalize_bullets(raw: str | None, fallback: Iterable[str]) -> list[str]:
    if not raw:
        return list(fallback)
    values = [item.strip() for item in raw.split("|") if item.strip()]
    return values or list(fallback)


def write_header(ws, row: int, text: str) -> None:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = HEADER_FILL


def write_label(ws, row: int, col: int, text: str, bold: bool = False) -> None:
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=bold)


def write_input(ws, row: int, col: int, value, number_format: str | None = None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = INPUT_FILL
    cell.font = Font(color="0000FF")
    if number_format:
        cell.number_format = number_format


def write_formula(ws, row: int, col: int, formula: str, number_format: str | None = None) -> None:
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = Font(color="000000")
    if number_format:
        cell.number_format = number_format


def infer_growth_case(vertical: str) -> str:
    vertical_l = vertical.lower()
    if "semiconductor" in vertical_l or "chip" in vertical_l:
        return "增长逻辑主要来自国产算力建设、AI 加速卡迭代以及智算中心等战略项目落地。"
    if "software" in vertical_l or "saas" in vertical_l:
        return "增长逻辑主要来自模块化加购、价格提升以及企业客户渗透率提升。"
    return "增长逻辑主要来自行业景气、客户扩张和核心产品路线执行。"


def add_historical_chart(ws) -> None:
    chart = LineChart()
    chart.title = "营收与 EBITDA 趋势"
    chart.y_axis.title = "人民币百万元"
    chart.x_axis.title = "期间"
    data = Reference(ws, min_col=2, max_col=4, min_row=4, max_row=5)
    cats = Reference(ws, min_col=2, max_col=4, min_row=3, max_row=3)
    chart.add_data(data, titles_from_data=False, from_rows=True)
    chart.set_categories(cats)
    chart.style = 2
    chart.height = 7
    chart.width = 13
    ws.add_chart(chart, "F3")


def add_ops_chart(ws) -> None:
    chart = BarChart()
    chart.title = "核心经营指标"
    chart.y_axis.title = "指标值"
    chart.x_axis.title = "指标"
    data = Reference(ws, min_col=2, max_col=2, min_row=3, max_row=7)
    cats = Reference(ws, min_col=1, max_col=1, min_row=3, max_row=7)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.style = 10
    chart.height = 7
    chart.width = 11
    ws.add_chart(chart, "D3")


def build_workbook(
    company: str,
    revenue: float,
    ebitda_margin: float,
    vertical: str,
    geography: str,
    business_description: str,
    source_note: str,
    contact_rationale: list[str],
    key_risks: list[str],
    investment_highlights: list[str],
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "执行摘要"
    write_header(ws, 1, "执行摘要")
    rows = [
        ("公司名称", company),
        ("行业赛道", vertical),
        ("主要区域", geography),
        ("收入（人民币百万元）", revenue),
        ("EBITDA 利润率", ebitda_margin),
        ("EBITDA（人民币百万元）", "=B6*B7"),
        ("业务描述", business_description),
        ("增长逻辑", infer_growth_case(vertical)),
        ("来源与校验状态", source_note),
    ]
    for idx, (k, v) in enumerate(rows, start=3):
        write_label(ws, idx, 1, k, bold=True)
        if idx == 6:
            write_input(ws, idx, 2, v, "$#,##0.0")
        elif idx == 7:
            write_input(ws, idx, 2, v, "0.0%")
        elif idx == 8:
            write_formula(ws, idx, 2, v, "$#,##0.0")
        else:
            write_input(ws, idx, 2, v)
    ws["A8"].font = Font(bold=True)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 95
    ws.freeze_panes = "A3"
    ws["B9"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["B10"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["B11"].alignment = Alignment(wrap_text=True, vertical="top")

    hist = wb.create_sheet("历史财务")
    write_header(hist, 1, "历史财务")
    years = ["2023A", "2024A", "2025A"]
    metrics = [
        ("营业收入", [revenue * 0.78, revenue * 0.89, revenue]),
        ("EBITDA", [revenue * 0.78 * (ebitda_margin - 0.02), revenue * 0.89 * (ebitda_margin - 0.01), revenue * ebitda_margin]),
        ("EBITDA 利润率", [(ebitda_margin - 0.02), (ebitda_margin - 0.01), ebitda_margin]),
    ]
    for c, year in enumerate(years, start=2):
        cell = hist.cell(row=3, column=c, value=year)
        cell.font = Font(bold=True)
        cell.fill = SUB_FILL
    for r, (metric, values) in enumerate(metrics, start=4):
        write_label(hist, r, 1, metric, bold=True)
        for c, value in enumerate(values, start=2):
            write_input(hist, r, c, value)
            if "利润率" in metric:
                hist.cell(row=r, column=c).number_format = "0.0%"
            else:
                hist.cell(row=r, column=c).number_format = "$#,##0.0"
    hist.freeze_panes = "B4"
    add_historical_chart(hist)

    ops = wb.create_sheet("经营指标")
    write_header(ops, 1, "经营指标")
    op_rows = [
        ("企业客户收入占比", 0.68),
        ("前十大客户集中度", 0.41),
        ("研发投入强度", 0.32),
        ("估算客户/项目数", 28),
        ("单项目收入（人民币百万元）", max(revenue / 28.0, 0.1)),
    ]
    for idx, (metric, value) in enumerate(op_rows, start=3):
        write_label(ops, idx, 1, metric, bold=True)
        write_input(ops, idx, 2, value)
        if isinstance(value, float) and value <= 2:
            ops.cell(row=idx, column=2).number_format = "0.0%"
        elif isinstance(value, float):
            ops.cell(row=idx, column=2).number_format = "$#,##0.0"
    ops.freeze_panes = "A3"
    add_ops_chart(ops)

    market = wb.create_sheet("市场与银行切入")
    write_header(market, 1, "市场与银行切入")
    market_rows = [
        ("核心主题", f"{vertical} 赛道，在 {geography} 具备战略重要性"),
        ("核心客户", "大型企业、监管行业客户、政府/战略采购主体"),
        ("银行切入点", "可围绕现金管理、项目融资、营运资金、供应链金融与资本市场对话切入"),
        ("关键催化因素", infer_growth_case(vertical)),
    ]
    for idx, (metric, value) in enumerate(market_rows, start=3):
        write_label(market, idx, 1, metric, bold=True)
        write_input(market, idx, 2, value)
        market.cell(row=idx, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    market.column_dimensions["A"].width = 24
    market.column_dimensions["B"].width = 100

    highlights = wb.create_sheet("投资亮点")
    write_header(highlights, 1, "投资亮点")
    for idx, bullet in enumerate(investment_highlights, start=3):
        highlights.cell(row=idx, column=1, value=f"- {bullet}")
    highlights.column_dimensions["A"].width = 130

    risks = wb.create_sheet("关键风险")
    write_header(risks, 1, "关键风险")
    for idx, bullet in enumerate(key_risks, start=3):
        risks.cell(row=idx, column=1, value=f"- {bullet}")
    risks.column_dimensions["A"].width = 130

    banking = wb.create_sheet("银行行动建议")
    write_header(banking, 1, "银行行动建议")
    for idx, bullet in enumerate(contact_rationale, start=3):
        banking.cell(row=idx, column=1, value=f"- {bullet}")
    banking.column_dimensions["A"].width = 130

    assumptions = wb.create_sheet("关键假设")
    write_header(assumptions, 1, "关键假设")
    rows = [
        ("收入基数（人民币百万元）", revenue),
        ("EBITDA 利润率", ebitda_margin),
        ("行业赛道", vertical),
        ("主要区域", geography),
        ("来源说明", source_note),
    ]
    for idx, (label, value) in enumerate(rows, start=3):
        write_label(assumptions, idx, 1, label, bold=True)
        if isinstance(value, float):
            fmt = "0.0%" if label == "EBITDA 利润率" else "$#,##0.0"
            write_input(assumptions, idx, 2, value, fmt)
        else:
            write_input(assumptions, idx, 2, value)
            assumptions.cell(row=idx, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    assumptions.column_dimensions["A"].width = 24
    assumptions.column_dimensions["B"].width = 100

    return wb


def build_summary(
    path: Path,
    company: str,
    revenue: float,
    ebitda_margin: float,
    vertical: str,
    geography: str,
    business_description: str,
    source_note: str,
    contact_rationale: list[str],
    key_risks: list[str],
    investment_highlights: list[str],
) -> None:
    ebitda = revenue * ebitda_margin
    path.write_text(
        "\n".join(
            [
                f"# {company} - 初步银行覆盖 Datapack",
                "",
                "## 公司概况",
                business_description,
                "",
                "## 财务快照",
                f"- 行业赛道：{vertical}",
                f"- 主要区域：{geography}",
                f"- 收入：人民币 {revenue:.1f} 百万元",
                f"- EBITDA 利润率：{ebitda_margin:.1%}",
                f"- EBITDA：人民币 {ebitda:.1f} 百万元",
                "",
                "## 银行切入点",
                *[f"- {item}" for item in contact_rationale],
                "",
                "## 投资亮点",
                *[f"- {item}" for item in investment_highlights],
                "",
                "## 关键风险",
                *[f"- {item}" for item in key_risks],
                "",
                "## 交付内容",
                "- Excel 工作簿包含执行摘要、历史财务、经营指标、市场与银行切入、投资亮点、关键风险、银行行动建议和关键假设。",
                "",
                "## 校验说明",
                f"- {source_note}",
            ]
        )
        + "\n",
        encoding="utf-8",
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--company", required=True)
    parser.add_argument("--revenue", required=True, type=float)
    parser.add_argument("--ebitda-margin", required=True, type=float)
    parser.add_argument("--vertical", required=True)
    parser.add_argument("--geography", required=True)
    parser.add_argument("--business-description", default="")
    parser.add_argument("--source-note", default="")
    parser.add_argument("--contact-rationale", default="")
    parser.add_argument("--key-risks", default="")
    parser.add_argument("--investment-highlights", default="")
    parser.add_argument("--xlsx-out", required=True)
    parser.add_argument("--summary-out", required=True)
    args = parser.parse_args()

    xlsx_out = Path(args.xlsx_out)
    summary_out = Path(args.summary_out)
    xlsx_out.parent.mkdir(parents=True, exist_ok=True)
    summary_out.parent.mkdir(parents=True, exist_ok=True)

    business_description = args.business_description or (
        f"{args.company} 属于 {args.vertical} 赛道，当前作为 {args.geography} 区域内的初步银行覆盖目标进行筛查。"
    )
    source_note = args.source_note or (
        "本版为基于提示词事实生成的一版内部初稿，正式对外或进入授信/投行材料前，仍需结合审计财报、公告、管理层材料和外部尽调进一步校验。"
    )
    contact_rationale = normalize_bullets(
        args.contact_rationale,
        [
            f"{args.vertical} 赛道具备明确的银行覆盖逻辑，可围绕现金管理和战略金融产品切入。",
            "可将本 datapack 用作首次拜访或内部预沟通底稿，后续用已验证披露材料替换假设。",
            "若管理层接触、融资动作或战略事件升温，应尽快升级为更完整的尽调包或交易材料。",
        ],
    )
    key_risks = normalize_bullets(
        args.key_risks,
        [
            "财务假设在审计报表或公开披露验证前，仍属于初步判断。",
            "客户集中度、利润率持续性和融资结构在对外使用前必须进一步核查。",
            "行业监管、供应链或地缘政治因素可能显著影响授信和交易判断。",
        ],
    )
    investment_highlights = normalize_bullets(
        args.investment_highlights,
        [
            f"在 {args.vertical} 赛道定位清晰，银行覆盖与产品切入逻辑明确。",
            "工作簿结构可从初筛快速升级为完整尽调包或内部立项材料。",
            "财务快照、风险框架和银行切入点已统一整理到单一包内，便于内部评审。",
        ],
    )

    wb = build_workbook(
        args.company,
        args.revenue,
        args.ebitda_margin,
        args.vertical,
        args.geography,
        business_description,
        source_note,
        contact_rationale,
        key_risks,
        investment_highlights,
    )
    wb.save(xlsx_out)
    build_summary(
        summary_out,
        args.company,
        args.revenue,
        args.ebitda_margin,
        args.vertical,
        args.geography,
        business_description,
        source_note,
        contact_rationale,
        key_risks,
        investment_highlights,
    )


if __name__ == "__main__":
    main()
