#!/usr/bin/env python3
"""
Build a minimal standalone LBO workbook with live Excel formulas.
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="F2F2F2")
OUTPUT_FILL = PatternFill("solid", fgColor="BDD7EE")
WHITE_BOLD = Font(name="Calibri", bold=True, color="FFFFFF")
BLACK_BOLD = Font(name="Calibri", bold=True, color="000000")
BLUE_FONT = Font(name="Calibri", color="0000FF")
BLACK_FONT = Font(name="Calibri", color="000000")
THIN = Side(style="thin", color="BFBFBF")


def border() -> Border:
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def set_input(ws, cell: str, value, source: str, number_format: str | None = None) -> None:
    ws[cell] = value
    ws[cell].fill = INPUT_FILL
    ws[cell].font = BLUE_FONT
    ws[cell].border = border()
    ws[cell].alignment = Alignment(horizontal="right")
    ws[cell].comment = Comment(source, "AIGroup")
    if number_format:
        ws[cell].number_format = number_format


def set_formula(ws, cell: str, formula: str, number_format: str | None = None, key_output: bool = False) -> None:
    ws[cell] = formula
    ws[cell].font = BLACK_BOLD if key_output else BLACK_FONT
    ws[cell].fill = OUTPUT_FILL if key_output else PatternFill(fill_type=None)
    ws[cell].border = border()
    ws[cell].alignment = Alignment(horizontal="right")
    if number_format:
        ws[cell].number_format = number_format


def set_label(ws, cell: str, text: str, bold: bool = False) -> None:
    ws[cell] = text
    ws[cell].font = BLACK_BOLD if bold else BLACK_FONT
    ws[cell].border = border()
    ws[cell].alignment = Alignment(horizontal="left")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", required=True)
    parser.add_argument("--company", default="ExampleCo")
    parser.add_argument("--currency", default="USD")
    parser.add_argument("--purchase-price", type=float, required=True)
    parser.add_argument("--revenue", type=float, required=True)
    parser.add_argument("--ebitda", type=float, required=True)
    parser.add_argument("--entry-multiple", type=float, required=True)
    parser.add_argument("--debt-pct", type=float, required=True)
    parser.add_argument("--cash-interest-rate", type=float, required=True)
    parser.add_argument("--exit-multiple", type=float, required=True)
    parser.add_argument("--ebitda-growth", required=True, help="Five comma-separated growth rates")
    args = parser.parse_args()

    growth = [float(x.strip()) for x in args.ebitda_growth.split(",") if x.strip()]
    if len(growth) != 5:
        raise SystemExit("ebitda-growth must contain exactly 5 comma-separated values")

    wb = Workbook()
    ws = wb.active
    ws.title = "LBO"

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.freeze_panes = "B5"

    ws["A1"] = f"{args.company} - Minimal LBO Model"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True)

    ws["A3"] = "Sources & Uses"
    ws["A3"].fill = HEADER_FILL
    ws["A3"].font = WHITE_BOLD
    ws.merge_cells("A3:B3")

    set_label(ws, "A4", "Purchase Price", True)
    set_input(ws, "B4", args.purchase_price, "Source: smoke test input", "$#,##0.0")
    set_label(ws, "A5", "Debt %", True)
    set_input(ws, "B5", args.debt_pct, "Source: smoke test input", "0.0%")
    set_label(ws, "A6", "Debt Financing", True)
    set_formula(ws, "B6", "=B4*B5", "$#,##0.0")
    set_label(ws, "A7", "Equity Contribution", True)
    set_formula(ws, "B7", "=B4-B6", "$#,##0.0", key_output=True)

    ws["D3"] = "Operating Assumptions"
    ws["D3"].fill = HEADER_FILL
    ws["D3"].font = WHITE_BOLD
    ws.merge_cells("D3:E3")

    set_label(ws, "D4", "Base Revenue", True)
    set_input(ws, "E4", args.revenue, "Source: smoke test input", "$#,##0.0")
    set_label(ws, "D5", "Base EBITDA", True)
    set_input(ws, "E5", args.ebitda, "Source: smoke test input", "$#,##0.0")
    set_label(ws, "D6", "Entry Multiple", True)
    set_input(ws, "E6", args.entry_multiple, "Source: smoke test input", '0.0"x"')
    set_label(ws, "D7", "Cash Interest Rate", True)
    set_input(ws, "E7", args.cash_interest_rate, "Source: smoke test input", "0.0%")
    set_label(ws, "D8", "Exit Multiple", True)
    set_input(ws, "E8", args.exit_multiple, "Source: smoke test input", '0.0"x"')

    ws["A10"] = "Projection & Returns"
    ws["A10"].fill = HEADER_FILL
    ws["A10"].font = WHITE_BOLD
    ws.merge_cells("A10:H10")

    headers = ["Metric", "LTM", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5", "Exit"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(11, idx)
        cell.value = header
        cell.fill = SECTION_FILL
        cell.font = BLACK_BOLD
        cell.border = border()
        cell.alignment = Alignment(horizontal="center")

    metrics = {
        12: "Revenue",
        13: "EBITDA",
        14: "EBITDA Margin",
        15: "Beginning Debt",
        16: "Cash Interest",
        17: "FCF Before Sweep",
        18: "Debt Paydown",
        19: "Ending Debt",
        20: "Exit Enterprise Value",
        21: "Exit Equity Value",
        22: "MoM",
    }
    for row, label in metrics.items():
        set_label(ws, f"A{row}", label, True)

    set_formula(ws, "B12", "=E4", "$#,##0.0")
    set_formula(ws, "B13", "=E5", "$#,##0.0")
    set_formula(ws, "B14", "=B13/B12", "0.0%")
    set_formula(ws, "B15", "=B6", "$#,##0.0")
    set_formula(ws, "B16", "=B15*$E$7", "$#,##0.0")
    set_formula(ws, "B17", "=B13-B16", "$#,##0.0")
    set_formula(ws, "B18", "=MIN(B17,B15)", "$#,##0.0")
    set_formula(ws, "B19", "=MAX(B15-B18,0)", "$#,##0.0")

    year_cols = ["C", "D", "E", "F", "G"]
    prior_cols = ["B", "C", "D", "E", "F"]
    for idx, col in enumerate(year_cols):
        prev = prior_cols[idx]
        growth_cell = f"{col}24"
        set_pct(ws, growth_cell, growth[idx], "Source: smoke test growth assumption")
        set_formula(ws, f"{col}12", f"={prev}12*(1+{growth_cell})", "$#,##0.0")
        set_formula(ws, f"{col}13", f"={prev}13*(1+{growth_cell})", "$#,##0.0")
        set_formula(ws, f"{col}14", f"={col}13/{col}12", "0.0%")
        set_formula(ws, f"{col}15", f"={prev}19", "$#,##0.0")
        set_formula(ws, f"{col}16", f"={col}15*$E$7", "$#,##0.0")
        set_formula(ws, f"{col}17", f"={col}13-{col}16", "$#,##0.0")
        set_formula(ws, f"{col}18", f"=MIN({col}17,{col}15)", "$#,##0.0")
        set_formula(ws, f"{col}19", f"=MAX({col}15-{col}18,0)", "$#,##0.0")

    set_formula(ws, "H20", "=G13*$E$8", "$#,##0.0", key_output=True)
    set_formula(ws, "H21", "=H20-G19", "$#,##0.0", key_output=True)
    set_formula(ws, "H22", "=H21/$B$7", '0.00"x"', key_output=True)

    ws["A24"] = "EBITDA Growth Assumptions"
    ws["A24"].fill = HEADER_FILL
    ws["A24"].font = WHITE_BOLD
    ws.merge_cells("A24:G24")

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def set_pct(ws, cell: str, value: float, source: str) -> None:
    set_input(ws, cell, value, source, "0.0%")


if __name__ == "__main__":
    main()
